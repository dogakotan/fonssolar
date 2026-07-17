#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fons Solar — PDF Servis
Excel dosyasini alir, gunluk_rapor_generator ile PDF uretir, geri doner.
Calistirmak: uvicorn main:app --port 8001 --reload
"""

import json, os, sys, tempfile
from pathlib import Path
from datetime import datetime

from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from gunluk_rapor_generator import generate_pdf

DEBUG_LOG = Path(__file__).with_name("pdf_debug.log")

app = FastAPI(title="Fons Solar PDF Service", version="1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],          # dev — prod'da domain ile sinirla
    allow_methods=["POST", "OPTIONS"],
    allow_headers=["*"],
)


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/generate-pdf")
async def pdf_endpoint(
    excel: UploadFile = File(..., description="Doldurulmus Excel raporu (.xlsx)"),
    proje_id: str = Form(default=None),
    tarih: str    = Form(default=None),
    proje_adi: str = Form(default=None),
    rapor_no: str = Form(default=None),
    hava: str = Form(default=None),
    hazirlayan: str = Form(default=None),
    photo_paths: str = Form(default=None),
):
    excel_path = None
    pdf_path   = None

    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(await excel.read())
            excel_path = tmp.name

        patch_excel_header(
            excel_path,
            proje_adi=proje_adi,
            tarih=tarih,
            rapor_no=rapor_no,
            hava=hava,
            hazirlayan=hazirlayan,
        )
        log_debug(
            "generate-pdf",
            proje_id=proje_id,
            tarih=tarih,
            proje_adi=proje_adi,
            rapor_no=rapor_no,
            hava=hava,
            hazirlayan=hazirlayan,
            photo_paths=photo_paths,
        )

        pdf_path = excel_path.replace(".xlsx", ".pdf")
        parsed_photo_paths = []
        if photo_paths:
            try:
                parsed_photo_paths = [p for p in json.loads(photo_paths) if isinstance(p, str) and p.strip()]
            except Exception:
                parsed_photo_paths = []

        generate_pdf(
            excel_path,
            output_path=pdf_path,
            proje_id=proje_id,
            tarih=tarih,
            storage_paths=parsed_photo_paths,
        )

        pdf_bytes = Path(pdf_path).read_bytes()
        fname = f"gunluk-rapor-{tarih or 'rapor'}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))

    finally:
        if excel_path and os.path.exists(excel_path):
            os.unlink(excel_path)
        if pdf_path and os.path.exists(pdf_path):
            os.unlink(pdf_path)


def patch_excel_header(
    excel_path: str,
    proje_adi: str = None,
    tarih: str = None,
    rapor_no: str = None,
    hava: str = None,
    hazirlayan: str = None,
):
    values = {
        "proje": proje_adi,
        "tarih": tarih,
        "rapor_no": rapor_no,
        "hava": hava,
        "hazirlayan": hazirlayan,
    }
    if not any(values.values()):
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    display_tarih = tarih
    if tarih and len(tarih) >= 10 and tarih[4:5] == "-" and tarih[7:8] == "-":
        display_tarih = f"{tarih[8:10]}.{tarih[5:7]}.{tarih[0:4]}"

    def put(cell, value):
        if not value:
            return
        for merged in list(ws.merged_cells.ranges):
            if cell in merged:
                ws.unmerge_cells(str(merged))
                break
        ws[cell] = value

    # Current template cells.
    put("B5", proje_adi)
    put("E5", display_tarih)
    put("H5", rapor_no)
    put("J5", hava)
    put("L5", hazirlayan)
    put("C114", hazirlayan)

    # Backward-compatible cells for older generator/service processes.
    put("C4", proje_adi)
    put("F4", display_tarih)
    put("I4", rapor_no)
    put("K4", hava)
    put("M4", hazirlayan)

    wb.save(excel_path)


def log_debug(event: str, **values):
    line = " | ".join([datetime.now().isoformat(timespec="seconds"), event] + [
        f"{key}={value!r}" for key, value in values.items()
    ])
    DEBUG_LOG.write_text(
        (DEBUG_LOG.read_text(encoding="utf-8") if DEBUG_LOG.exists() else "") + line + "\n",
        encoding="utf-8",
    )
